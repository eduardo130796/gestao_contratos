import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO


# Inicializar session_state para o login
if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False


CONFIG_FILE = "config.json"
AUTH_PASSWORD = "Eduardo13"  # Defina a senha para acessar as configurações

def atualizar_planilha_streamlit(arquivo_base, arquivo_atualizacao):
# Abre a planilha base com openpyxl
    wb = openpyxl.load_workbook(arquivo_base)
    ws = wb.active  # Aba principal da planilha base

    # Carrega a planilha de atualização e a de contratos com pandas
    df_atualizacao = pd.read_excel(arquivo_atualizacao, skiprows=2)


    # --- Atualiza Valor Empenhado ---
    # Criar mapa {nota de empenho (ult 8 digitos): valor empenhado}
    mapa_valor_empenhado = {
        str(row["Número da Nota de Empenho"]).strip()[-8:]: row["Saldo - R$ (Item Informação)"]
        for _, row in df_atualizacao.iterrows()
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[1]  # Coluna B: Nota de Empenho
        valor_empenhado_cell = row[5]  # Coluna F: Valor Empenhado

        nota_empenho = str(nota_empenho_cell.value).strip()
        if nota_empenho in mapa_valor_empenhado:
            novo_valor = mapa_valor_empenhado[nota_empenho]
            if valor_empenhado_cell.value != novo_valor:
                print(f"Atualizando valor empenhado da NE {nota_empenho}: {valor_empenhado_cell.value} -> {novo_valor}")
                valor_empenhado_cell.value = novo_valor


    # Atualização de pagamentos
    df_atualizacao = df_atualizacao[~df_atualizacao.apply(lambda row: row.astype(str).str.contains('Total').any(), axis=1)]
    df_atualizacao.ffill(inplace=True)

    # --- Atualiza pagamentos mês a mês ---
    meses = {
        "jan": 7,  "fev": 8,  "mar": 9,  "abr": 10, "mai": 11, "jun": 12,
        "jul": 13, "ago": 14, "set": 15, "out": 16, "nov": 17, "dez": 18
    }

    meses_ingles_para_portugues = {
        'jan': 'jan', 'feb': 'fev', 'mar': 'mar', 'apr': 'abr', 'may': 'mai',
        'jun': 'jun', 'jul': 'jul', 'aug': 'ago', 'sep': 'set', 'oct': 'out',
        'nov': 'nov', 'dec': 'dez'
    }

    # Agrupa pagamentos por nota e mês
    pagamentos_por_nota = {}
    for _, row in df_atualizacao.iterrows():
        nota_empenho = str(row["Número da Nota de Empenho"]).strip()[-8:]
        data_pagamento = row["Métrica"]
        valor_pago = row["Unnamed: 13"]

        if nota_empenho not in pagamentos_por_nota:
            pagamentos_por_nota[nota_empenho] = {mes: 0.0 for mes in meses}

        # Converter data pagamento para datetime e obter mês em pt-br
        data_pagamento = pd.to_datetime(data_pagamento, errors='coerce', dayfirst=True)
        if pd.notna(data_pagamento):
            mes_abrev = data_pagamento.strftime('%b').lower()
            mes_port = meses_ingles_para_portugues.get(mes_abrev)
            if mes_port in meses:
                pagamentos_por_nota[nota_empenho][mes_port] += float(valor_pago) if not pd.isna(valor_pago) else 0.0

    # Atualiza células dos meses
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[1]  # Coluna B: Nota de Empenho
        nota_empenho = str(nota_empenho_cell.value).strip()

        if nota_empenho in pagamentos_por_nota:
            for mes, idx_col in meses.items():
                valor_mes = pagamentos_por_nota[nota_empenho].get(mes, 0.0)
                cell_mes = row[idx_col]
                if cell_mes.value != valor_mes:
                    print(f"Atualizando pagamento {mes} da NE {nota_empenho}: {cell_mes.value} -> {valor_mes}")
                    cell_mes.value = valor_mes

    # Atualiza células dos meses e calcula valor pago real
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[1]  # Coluna B
        nota_empenho = str(nota_empenho_cell.value).strip()

        if nota_empenho in pagamentos_por_nota:
            total_pago_real = 0.0
            for mes, idx_col in meses.items():
                valor_mes = pagamentos_por_nota[nota_empenho].get(mes, 0.0)
                cell_mes = row[idx_col]
                if cell_mes.value != valor_mes:
                    print(f"Atualizando pagamento {mes} da NE {nota_empenho}: {cell_mes.value} -> {valor_mes}")
                    cell_mes.value = valor_mes
                total_pago_real += valor_mes

            # Atualiza valor pago (coluna G = índice 6)
            valor_pago_cell = row[6]
            if valor_pago_cell.value != total_pago_real:
                print(f"Atualizando valor pago da NE {nota_empenho}: {valor_pago_cell.value} -> {total_pago_real}")
                valor_pago_cell.value = total_pago_real

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.title("🔄 Atualizar Planilha de Notas")

# Se o usuário não está autenticado, pede senha
if not st.session_state["autenticado"]:
    senha = st.text_input("Digite a senha para acessar:", type="password")
    if st.button("🔑 Entrar"):
        if senha == AUTH_PASSWORD:
            st.session_state["autenticado"] = True
            st.success("✅ Acesso concedido!")
            st.rerun()
        else:
            st.error("❌ Senha incorreta! Tente novamente.")

# Se já está autenticado, mostra configurações
if st.session_state["autenticado"]:

    col1, col2 = st.columns(2)
    with col1:
        planilha_base = st.file_uploader("📤 Envie a Planilha de Notas (base)", type="xlsx")
    with col2:
        planilha_atualizacao = st.file_uploader("📥 Envie a Planilha de Atualização", type="xlsx")

    if planilha_base and planilha_atualizacao:
        if st.button("🚀 Atualizar Planilha"):
            arquivo_atualizado = atualizar_planilha_streamlit(planilha_base, planilha_atualizacao)
            st.success("✅ Planilha atualizada com sucesso!")
            st.download_button("⬇️ Baixar Planilha Atualizada", data=arquivo_atualizado, file_name="planilha_notas_atualizada.xlsx")
    # Botão de logout
    if st.button("🚪 Sair"):
        st.session_state["autenticado"] = False
        st.rerun()        